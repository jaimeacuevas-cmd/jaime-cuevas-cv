"""Excel ingestion module - reads CV_Dataset_Maestro_Jaime_Cuevas.xlsx."""

import logging
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelReader:
    """Reads and parses Excel master file."""

    # Mapping: Sheet name → Entity type
    SHEET_TYPE_MAP = {
        'Agentes_y_Artistas': 'Agent',
        'Organizaciones': 'Organization',
        'Lugares_y_Sedes': 'Location',
        'Tesauro_SKOS': 'Concept',
        'Formacion_Academica': 'Education',
        'Trayectoria_Laboral': 'Position',
        'Publicaciones': 'Publication',
        'Exposiciones_Curadurias': 'Exhibition',
        'Proyectos_y_Fondos': 'Project',
        'Medios_y_Congresos': 'Media',
        'Portafolio_Digital_Web': 'DigitalHeritage',
        'Relaciones_Grafo_LOD': 'Link',
    }

    def __init__(self, excel_path: str):
        """Initialize Excel reader with path to master file."""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        logger.info(f"Initialized ExcelReader for: {self.excel_path}")

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Read single sheet from Excel.
        Returns DataFrame with minimal cleaning.
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            logger.info(f"Read sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
            return df
        except Exception as e:
            logger.error(f"Failed to read sheet '{sheet_name}': {e}")
            raise

    def read_all_sheets(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Read all entity sheets (excludes Relaciones_Grafo_LOD for now).
        Returns dict: {'Agent': [...rows...], 'Organization': [...], ...}
        """
        result = {}

        for sheet_name, entity_type in self.SHEET_TYPE_MAP.items():
            if entity_type == 'Link':
                # Handle relations separately
                continue

            try:
                df = self.read_sheet(sheet_name)

                # Convert to list of dicts
                rows = df.fillna('').to_dict('records')

                # Filter out completely empty rows
                rows = [r for r in rows if any(v for v in r.values() if v != '')]

                if entity_type not in result:
                    result[entity_type] = []

                result[entity_type].extend(rows)
                logger.info(f"  → {entity_type}: {len(rows)} valid rows")

            except Exception as e:
                logger.warning(f"Error reading sheet '{sheet_name}': {e}")
                continue

        return result

    def read_relations(self) -> List[Dict[str, Any]]:
        """
        Read Relaciones_Grafo_LOD sheet.
        Returns list of relationship records.
        """
        try:
            df = self.read_sheet('Relaciones_Grafo_LOD')
            rows = df.fillna('').to_dict('records')
            rows = [r for r in rows if any(v for v in r.values() if v != '')]

            logger.info(f"Read relations: {len(rows)} valid records")
            return rows

        except Exception as e:
            logger.warning(f"Error reading Relaciones_Grafo_LOD: {e}")
            return []

    def get_sheet_info(self) -> Dict[str, Any]:
        """
        Get metadata about all sheets in workbook.
        Returns dict with sheet names, row counts, etc.
        """
        try:
            wb = load_workbook(self.excel_path)
            info = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                info[sheet_name] = {
                    'row_count': ws.max_row,
                    'column_count': ws.max_column,
                    'entity_type': self.SHEET_TYPE_MAP.get(sheet_name, 'Unknown'),
                }

            return info

        except Exception as e:
            logger.warning(f"Error reading workbook info: {e}")
            return {}
